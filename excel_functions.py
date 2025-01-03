from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.chart import LineChart, Reference
from openpyxl.chart import BarChart, Reference
from openpyxl.chart import AreaChart, Reference
import openpyxl.utils
from pathlib import Path
import requests
from main import regex, temperatura_de_lugar, icono_clima, name_icono
from datetime import datetime
import os


API_KEY = 'lkzersr1hytpof1wj54ma832l6nr1wuqshjjx12a'
BASE_URL = "https://www.meteosource.com/api/v1/free/point"

def obtener_datos_clima(place):
    try:
        parameters = {'key': API_KEY, 'place_id': place, 'units': 'metric'}
        response = requests.get(BASE_URL, params=parameters)
        response.raise_for_status() 
        
        data = response.json()
        hourly_data = data.get("hourly", {}).get("data", [])
        clima_datos = []
        
        for entry in hourly_data[:12]: 
            time = entry.get("date")
            temp = entry.get('temperature')
            wind_speed = entry.get('wind', {}).get('speed', 0) * 3.6 
            clima_datos.append((time, temp, wind_speed))
        
        return clima_datos
    except Exception as e:
        print("Error al obtener los datos:", e)
        return []

def guardar_en_excel(place):
    if not regex(place):
        print("El lugar ingresado no es válido. Debe contener solo letras.")
        return
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Datos del Clima"
    
    temp_actual = temperatura_de_lugar(place)

    ws.merge_cells('A1:D1')  
    temp_cell = ws['A1']
    temp_cell.value = f"Temperatura Actual: {temp_actual} °C"
    temp_cell.font = Font(size=24, bold=True, color="FFFFFF")  
    temp_cell.fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid") 
    temp_cell.alignment = Alignment(horizontal="center", vertical="center") 
    
    headers = ["Hora", "Temperatura (°C)", "Velocidad Viento (km/h)", "Ícono", "Nombre Ícono"]
    ws.append(headers)
    
    clima_datos = obtener_datos_clima(place)
    
    for dato in clima_datos:
        time_obj = datetime.fromisoformat(dato[0])
        time_str = time_obj.strftime("%H:%M")
        temp = dato[1]
        wind_speed = dato[2]
        icon_url = icono_clima(place)
        icon_name = name_icono(place)
        
        ws.append([time_str, temp, wind_speed, icon_url, icon_name])
    
    for cell in ws[2]:
        cell.font = Font(bold=True, color="FFFFFF") 
        cell.fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")  
    
    column_widths = [12, 20, 25, 40, 30]  
    for i, width in enumerate(column_widths, start=1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width

    '''Se define la primer grafica y se establecen sus valores
    chart = LineChart()
    chart.title = "Temperatura y Velocidad del Viento"
    chart.style = 13
    chart.x_axis.title = "Hora"
    chart.y_axis.title = "Valores".'''
    chart = LineChart()
    chart.title = "Temperatura y Velocidad del Viento"
    chart.style = 13
    chart.x_axis.title = "Hora"
    chart.y_axis.title = "Valores"
    '''Se establecen el rango de columnas y filas de cuales tomara los datos para la grafica
    temp_series = Reference(ws, min_col=2, min_row=2, max_row=len(clima_datos) + 2)
    wind_series = Reference(ws, min_col=3, min_row=2, max_row=len(clima_datos) + 2).'''
    temp_series = Reference(ws, min_col=2, min_row=2, max_row=len(clima_datos) + 2)
    wind_series = Reference(ws, min_col=3, min_row=2, max_row=len(clima_datos) + 2)
    '''Agregar los datos a la grafica.'''
    chart.add_data(temp_series, titles_from_data=True)
    chart.add_data(wind_series, titles_from_data=True)

    chart.series[0].graphicalProperties.line.solidFill = "FF0000"  
    chart.series[1].graphicalProperties.line.solidFill = "00FF00" 

    horas = Reference(ws, min_col=1, min_row=3, max_row=len(clima_datos) + 2)
    chart.set_categories(horas)
    '''Los muestra y agrega en el excel empezando la grafica en la celda G5.'''
    ws.add_chart(chart, "G5")
    
    

    '''Segunda grafica Barras
    chart= BarChart()
    chart.title = "temperatura durante la utlima hora"
    chart.style=10
    chart.x_axis.title = "hora"
    chart.y_axis.title = "temperaturas".'''
    chart= BarChart()
    chart.title = "temperatura durante la utlima hora"
    chart.style=10
    chart.x_axis.title = "hora"
    chart.y_axis.title = "temperaturas"
    '''Se establecen el rango de columnas y filas de cuales tomara los datos para la grafica
    hora_series = Reference(ws, min_col=1, min_row=2, max_row=len(clima_datos) + 2)
    temp_series = Reference(ws, min_col=2, min_row=2, max_row=len(clima_datos) + 2).'''
    hora_series = Reference(ws, min_col=1, min_row=2, max_row=len(clima_datos) + 2)
    temp_series = Reference(ws, min_col=2, min_row=2, max_row=len(clima_datos) + 2)
    '''Agregar los datos a la grafica
    chart.add_data(hora_series, titles_from_data=True)
    chart.set_categories(temp_series).'''
    chart.add_data(hora_series, titles_from_data=True)
    chart.set_categories(temp_series)
    '''Los muestra y agrega en el excel empezando la grafica en la celda G20.'''
    ws.add_chart(chart, "G20")
    '''Tercera grafica Area.'''
    chart=AreaChart()
    chart.title = "vientos por hora"
    chart.style = 9
    chart.x_axis.title="hora"
    chart.y_axis.title="vientos (km/h)"

    hora_series = Reference(ws, min_col=1, min_row=2, max_row=len(clima_datos) +2)
    vientos_series = Reference(ws, min_col=3, min_row=2, max_row=len(clima_datos) +2)

    chart.set_categories(hora_series)
    chart.add_data(vientos_series, titles_from_data=True)

    ws.add_chart(chart, "G35")
    '''Se guarda el archivo empezando con el nombre que se consulto seguido del año, mes y hora de consulta
    nombre_archivo = Path(f"{place}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx").'''
    nombre_archivo = Path(f"{place}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx")
    '''Nombre archivo que contiene la ruta del archivo se guarda como workbook
    wb.save(nombre_archivo).'''
    try:
        wb.save(nombre_archivo)
        '''Se imprime la salida de la ruta del archivo
        print(f"Archivo guardado en: {nombre_archivo}").'''
        print(f"Archivo guardado en: {nombre_archivo}")
        '''Con el modulo os abrimos automaticamente el archivo en excel despues de que este se guarde
        os.startfile(nombre_archivo).'''
        os.startfile(nombre_archivo)
        
    except PermissionError:
        print(f"No se pudo guardar el archivo. Asegúrate de que el archivo no esté abierto y que tienes permisos de escritura en {nombre_archivo}.")

'''if __name__ == "__main__":
    lugar = "monterrey" 
    print("Temperatura actual:", temperatura_de_lugar(lugar), "°C")
    print("Ícono del clima:", icono_clima(lugar))
    print("Nombre del ícono:", name_icono(lugar))
    print("Hora:", time_zone(lugar))
    print("Previsión del viento:", forecast(lugar))
    guardar_en_excel(lugar)
'''
