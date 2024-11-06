import requests
from datetime import datetime
import re
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.chart import LineChart, Reference
import openpyxl.utils
from pathlib import Path


API_KEY = 'rxsjl8vt5wbc5x5twoz149tf8bfnbgzsm1c0vcnb'
BASE_URL = "https://www.meteosource.com/api/v1/free/point"

def regex(place):
    exp = r"^[a-zA-ZáéíóúÁÉÍÓÚñÑ]+(\s+[a-zA-ZáéíóúÁÉÍÓÚñ]+)*$"
    matches = re.match(exp, place)
    return matches is not None 

def obtener_datos_clima(place):
    try:
        parameters = {'key': API_KEY, 'place_id': place, 'units': 'metric'}
        response = requests.get(BASE_URL, params=parameters)
        response.raise_for_status() 
        
        data = response.json()
        hourly_data = data.get("hourly", {}).get("data", [])
        clima_datos = []
        
        for entry in hourly_data[:12]: 
            time = entry.get("date")
            temp = entry.get('temperature')
            wind_speed = entry.get('wind', {}).get('speed', 0) * 3.6 
            clima_datos.append((time, temp, wind_speed))
        
        return clima_datos
    except Exception as e:
        print("Error al obtener los datos:", e)
        return []

def temperatura_de_lugar(place):
    try:
        parameters = {'key': API_KEY, 'place_id': place, 'units': 'metric'}
        response = requests.get(BASE_URL, params=parameters)
        response.raise_for_status()
        data = response.json()
        temp = data['current']['temperature']
        return temp 
    except Exception as e:
        print("Lugar geográfico no accesible, intente de nuevo:", e)

def icono_clima(place):
    try:
        parameters = {'key': API_KEY, 'place_id': place}
        response = requests.get(BASE_URL, params=parameters)
        response.raise_for_status()
        data = response.json()
        icono = data['current']['icon_num']
        return f"https://www.meteosource.com/static/img/ico/weather/{icono}.svg"
    except Exception as e:
        print('Imagen no disponible:', e)
        
def name_icono(place):
    try:
        parameters = {'key': API_KEY, 'place_id': place}
        response = requests.get(BASE_URL, params=parameters)
        response.raise_for_status()
        data = response.json()
        icon_name = data['current']['summary']
        return f"{icon_name}"
    except Exception as e:
        return f'No se encuentra el estado actual del clima en {place}: {e}'
    
def time_zone(place):
    params = {'key': API_KEY, 'place_id': place}
    try:
        response = requests.get(BASE_URL, params=params)
        response.raise_for_status()
        data = response.json()
        hour_str = data["hourly"]["data"][0]["date"]
        hour_obj = datetime.fromisoformat(hour_str)
        return hour_obj.strftime("%H:%M")
    except Exception as e:
        return f"Hora no disponible: {e}" 

def forecast(place):
    params= {'key': API_KEY, 'place_id': place, 'units': 'metric'} 
    try:
        response = requests.get(BASE_URL, params=params)
        response.raise_for_status()
        data = response.json()
        wind_speed = data['current']['wind']['speed']
        conv = wind_speed * 3.6
        wind_dir = data['current']['wind']['dir']
        return f'{conv:.2f} km/h {wind_dir}'
    except Exception as e:
        return f"Error al obtener la previsión: {e}"

def guardar_en_excel(place):
    if not regex(place):
        print("El lugar ingresado no es válido. Debe contener solo letras.")
        return
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Datos del Clima"
    
    temp_actual = temperatura_de_lugar(place)

    ws.merge_cells('A1:D1')  
    temp_cell = ws['A1']
    temp_cell.value = f"Temperatura Actual: {temp_actual} °C"
    temp_cell.font = Font(size=24, bold=True, color="FFFFFF")  
    temp_cell.fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid") 
    temp_cell.alignment = Alignment(horizontal="center", vertical="center") 
    
    headers = ["Hora", "Temperatura (°C)", "Velocidad Viento (km/h)", "Ícono", "Nombre Ícono"]
    ws.append(headers)
    
    clima_datos = obtener_datos_clima(place)
    
    for dato in clima_datos:
        time_obj = datetime.fromisoformat(dato[0])
        time_str = time_obj.strftime("%H:%M")
        temp = dato[1]
        wind_speed = dato[2]
        icon_url = icono_clima(place)
        icon_name = name_icono(place)
        
        ws.append([time_str, temp, wind_speed, icon_url, icon_name])
    
    for cell in ws[2]:
        cell.font = Font(bold=True, color="FFFFFF") 
        cell.fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")  
    
    column_widths = [12, 20, 25, 40, 30]  
    for i, width in enumerate(column_widths, start=1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width
    
    chart = LineChart()
    chart.title = "Temperatura y Velocidad del Viento"
    chart.style = 13
    chart.x_axis.title = "Hora"
    chart.y_axis.title = "Valores"

    temp_series = Reference(ws, min_col=2, min_row=2, max_row=len(clima_datos) + 2)
    wind_series = Reference(ws, min_col=3, min_row=2, max_row=len(clima_datos) + 2)

    chart.add_data(temp_series, titles_from_data=True)
    chart.add_data(wind_series, titles_from_data=True)

    chart.series[0].graphicalProperties.line.solidFill = "FF0000"  
    chart.series[1].graphicalProperties.line.solidFill = "00FF00" 

    horas = Reference(ws, min_col=1, min_row=3, max_row=len(clima_datos) + 2)
    chart.set_categories(horas)
    
    ws.add_chart(chart, "G5")
    
    nombre_archivo = Path({place}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx")
    
    try:
        wb.save(nombre_archivo)
        print(f"Archivo guardado en: {nombre_archivo}")
        with open(nombre_archivo,'rb') as file:
            content=file.read()
    except PermissionError:
        print(f"No se pudo guardar el archivo. Asegúrate de que el archivo no esté abierto y que tienes permisos de escritura en {nombre_archivo}.")

if __name__ == "__main__":
    lugar = "monterrey" 
    print("Temperatura actual:", temperatura_de_lugar(lugar), "°C")
    print("Ícono del clima:", icono_clima(lugar))
    print("Nombre del ícono:", name_icono(lugar))
    print("Hora:", time_zone(lugar))
    print("Previsión del viento:", forecast(lugar))
    guardar_en_excel(lugar)
